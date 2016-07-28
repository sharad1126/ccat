
import openpyxl
import os
import collections
import sys
from openpyxl.styles import Font

wb = openpyxl.load_workbook('Sample.xlsx')
sheets = wb.get_sheet_names()
print sheets[0]
sheetName = sheets[0]
workSheet = wb.get_sheet_by_name(sheetName)


wb_output = openpyxl.Workbook()
workSheet_output = wb_output.active
workSheet_output.title = "Cdr-Report"
fontObj1 = Font(name='Times New Roman', bold=True, size=14)
workSheet_output.cell(row = 1,column = 1).value = 'IMEI'



workSheet_output.cell(row = 1,column = 1).font = fontObj1
workSheet_output.cell(row = 3,column = 1).font = fontObj1
workSheet_output.cell(row = 4,column = 4).font = fontObj1
workSheet_output.cell(row = 4,column = 6).font = fontObj1
workSheet_output.cell(row = 4,column = 8).font = fontObj1
workSheet_output.cell(row = 4,column = 10).font = fontObj1
workSheet_output.cell(row = 4,column = 12).font = fontObj1
workSheet_output.cell(row = 4,column = 14).font = fontObj1


workSheet_output.column_dimensions['A'].width = 12
workSheet_output.column_dimensions['D'].width = 27
workSheet_output.column_dimensions['F'].width = 27
workSheet_output.column_dimensions['H'].width = 25
workSheet_output.column_dimensions['J'].width = 25
workSheet_output.column_dimensions['L'].width = 26
workSheet_output.column_dimensions['N'].width = 26












TotalEntries = workSheet.max_row
print '"\n Total Number of available Entries in CDR are :',TotalEntries 

count_incoming_call=0
count_outgoing_call=0
count_incoming_sms=0
count_outgoing_sms=0
for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="IN___CALL"):
		count_incoming_call=count_incoming_call+1
		

print "\n Total Number of Incoming calls:",count_incoming_call
	

for i in range(1,TotalEntries):
	 if((workSheet.cell(row=i,column=8)).value=="OUT_CALL"):
		count_outgoing_call=count_outgoing_call+1

print "\n Total Number of Outgoing Calls:",count_outgoing_call

for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="IN___SMS"):
		count_incoming_sms=count_incoming_sms+1

print "\n Total Number of Incoming Sms:",count_incoming_sms

for i in range(1,TotalEntries):
	 if((workSheet.cell(row=i,column=8)).value=="OUT_SMS"):
		count_outgoing_sms=count_outgoing_sms+1

print "\n Total Number of Outgoing SMS:",count_outgoing_sms
print "----------------------------------------------------------------------------------------------"
set_calledNumbers=set()
for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="OUT_CALL"):
		calledNumber=workSheet.cell(row=i,column=2).value
		if calledNumber not in set_calledNumbers:
			set_calledNumbers.add(calledNumber)


print "\n Unique Called Numbers are :"
called_count=0
set_calledNumbers = list(set_calledNumbers)
for elem in (set_calledNumbers):
	called_count=called_count+1
	print "\n ",called_count,"=",elem

wb_output = openpyxl.Workbook()
workSheet_output = wb_output.active
workSheet_output.title = "CdrAnalysis Report"
fontObj1 = Font(name = 'Times New Roman',bold=True,size=14)
workSheet_output.cell(row=1,column=1).value = "IMEI"
workSheet_output.cell(row=3,column=1).value = "IMSI" 
workSheet_output.cell(row=4,column=4).value = "FREQUENTLY CALLED NUMBERS"
workSheet_output.cell(row=4,column=6).value = "FREQUENTLY RECEIVED NUMBERS"
workSheet_output.cell(row=4,column=8).value = "UNIQUE CALLED NUMBERS"
workSheet_output.cell(row=4,column=10).value = "UNIQUE RECEIVED NUMBERS"
workSheet_output.cell(row=4,column=12).value = "SMS OUTGOING FREQUENT"
workSheet_output.cell(row=4,column=14).value = "SMS INCOMING FREQUENT"







print "------------------------------------------------------------------------------------------------"


c = workSheet.max_column
r = workSheet.max_row

k = 0 

for i in range(2, r+1):
	if k != workSheet.cell(row = i+2, column = 3).value:
		k = workSheet.cell(row =i+2, column =3).value
		print k
		print "First Call/SMS"
		if workSheet.cell(row = i+2, column = 8).value == 'IN___CALL':
			print "Incoming Call from : " + workSheet.cell(row = i+2, column = 1).value
		elif workSheet.cell(row = i+2, column = 8).value == 'IN___SMS':	
			print "Incoming SMS from : " + workSheet.cell(row = i+2, column = 1).value
		elif workSheet.cell(row = i+2, column = 8).value == 'OUT___CALL':
			print "Outgoing Call to : " + workSheet.cell(row = i+2, column = 2).value
		#else:
			#print "Outgoing SMS to : " + workSheet.cell(row = i+2, column = 2).value
		print "Last Call/SMS"
		if workSheet.cell(row = i+1, column = 8).value == 'IN___CALL':
			print "Incoming Call from : " + workSheet.cell(row = i+1, column = 1).value
		elif workSheet.cell(row = i+1, column = 8).value == 'IN___SMS':	
			print "Incoming SMS from : " + workSheet.cell(row = i+1, column = 1).value
		elif workSheet.cell(row = i+1, column = 8).value == 'OUT___CALL':
			print "Outgoing Call to : " + workSheet.cell(row = i+1, column = 2).value
		else:
			print workSheet.cell(row = i+1, column = 2).value == 'OUT_SMS'

Final_IMEI = 0
print "ESN/IMEI No. : "

for i in range(1, r+1):
	for j in range(1, c+1):
		if workSheet.cell(row = i, column = j).value == 'ESN_or_IMEI_NO':
			print workSheet.cell(row = i+2, column = j).value
			Final_IMEI = workSheet.cell(row=i+2 , column=j).value

workSheet_output.cell(row=1,column=2).value = Final_IMEI
			
Final_IMSI = 0
print "MIN/IMSI No. : "
			
for i in range(1, r+1):
	for j in range(1, c+1):
		if workSheet.cell(row = i, column = j).value == 'MIN_or_IMSI_NO':
			print workSheet.cell(row = i+2, column = j).value
			workSheet.cell(row = i+2, column = j).value = Final_IMSI

workSheet_output.cell(row=3,column=2).value = Final_IMSI	



set_callingNumbers=set()	
for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="IN___CALL"):
		callingNumber=workSheet.cell(row=i,column=1).value
		if callingNumber not in set_callingNumbers:
			set_callingNumbers.add(callingNumber)

print "\n Unique Recieved Called Numbers :"
count_received = 0
set_callingNumbers = list(set_callingNumbers)
for elem in set_callingNumbers:
	count_received = count_received + 1
	print "\n ",count_received,"-",elem
temp_counter=0
for elem in set_callingNumbers:
	temp_counter = temp_counter+1
	workSheet_output.cell(row=temp_counter+4,column=10).value = elem
	

print "--------------------------------------------------------------------------------------------------"

List_OutGoingNumbers=list()
for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="OUT_CALL"):
		outgoingNumber=workSheet.cell(row=i,column=2).value
		List_OutGoingNumbers.append(outgoingNumber)
counterOutgoing = 0

for elem in List_OutGoingNumbers :
	counterOutgoing = counterOutgoing+1
	workSheet_output.cell(row = counterOutgoing+4,column=8).value = elem



	

counter_outgoing_call=collections.Counter(List_OutGoingNumbers)
print(counter_outgoing_call)
print "\n Most Frequently Called Number:",counter_outgoing_call.most_common(3)
counter_frequently = 0
for elem in counter_outgoing_call.most_common(3):
	counter_frequently = counter_frequently + 1
	print "\n ",counter_frequently,"-",elem
local_outgoing_counter = 0
for elem in counter_outgoing_call.most_common(3):
	local_outgoing_counter = local_outgoing_counter + 1
	workSheet_output.cell(row=local_outgoing_counter+4,column=4).value=float(elem[0])




	


print "---------------------------------------------------------------------------------------------------"

List_Incoming_number=list()
for i in range(1,TotalEntries):
	if((workSheet.cell(row=i,column=8)).value=="IN___CALL"):
		incoming_calling_number=workSheet.cell(row=i,column=1).value
		List_Incoming_number.append(incoming_calling_number)

counter_incoming_call=collections.Counter(List_Incoming_number)
print(counter_incoming_call)
print "Most Frequently Received Called Numbers:",counter_incoming_call.most_common(3)
local_incoming_counter=0
for elem in counter_incoming_call.most_common(3):
	local_incoming_counter = local_incoming_counter + 1
	workSheet_output.cell(row=local_incoming_counter+4,column=6).value=float(elem[0])

print "-----------------------------------------------------------------------------------------------------"
List_IncomingSms=list()
for i in range(1,TotalEntries):
	 if((workSheet.cell(row=i,column=8)).value=="IN___SMS"):
		incoming_sms_number=workSheet.cell(row=i,column=1).value
		List_IncomingSms.append(incoming_sms_number)

counter_incoming_sms=collections.Counter(List_IncomingSms)
print(counter_incoming_sms)
print "\n Most Frequently Incoming SMS Numbers:",counter_incoming_sms.most_common(3)
counter_incoming = 0
for elem in counter_incoming_sms.most_common(3):
	counter_incoming = counter_incoming + 1
	print "\n ",counter_incoming , "-", elem




		
print "-----------------------------------------------------------------------------------------------------"
List_Outgoing_sms=list()
for i in range(1,TotalEntries):
	 if((workSheet.cell(row=i,column=8)).value=="OUT_SMS"):
		outgoing_sms_number=workSheet.cell(row=i,column=2).value
		List_Outgoing_sms.append(outgoing_sms_number)

counter_outgoing_sms=collections.Counter(List_Outgoing_sms)
print(counter_outgoing_sms)
print "\n Most Frequently Outgoing SMS Numbers:",counter_outgoing_sms.most_common(3)
counter_outgoing = 0
for elem in counter_outgoing_sms.most_common(3):
	counter_outgoing = counter_outgoing+1
	print "\n ",counter_outgoing,"-",elem






wb_output.save('Cdr-Report-Final-2.xlsx')

import openpyxl
wb = openpyxl.load_workbook('Sample.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('sheet1')
c = sheet.max_column
r = sheet.max_row

k = 0 

for i in range(2, r+1):
	if k != sheet.cell(row = i+2, column = 3).value:
		k = sheet.cell(row =i+2, column =3).value
		print k
		print "First Call/SMS"
		if sheet.cell(row = i+2, column = 8).value == 'IN___CALL':
			print "Incoming Call from : " + sheet.cell(row = i+2, column = 1).value
		elif sheet.cell(row = i+2, column = 8).value == 'IN___SMS':	
			print "Incoming SMS from : " + sheet.cell(row = i+2, column = 1).value
		elif sheet.cell(row = i+2, column = 8).value == 'OUT___CALL':
			print "Outgoing Call to : " + sheet.cell(row = i+2, column = 2).value
		else:
			print "Outgoing SMS to : " + sheet.cell(row = i+2, column = 2).value
		print "Last Call/SMS"
		if sheet.cell(row = i+1, column = 8).value == 'IN___CALL':
			print "Incoming Call from : " + sheet.cell(row = i+1, column = 1).value
		elif sheet.cell(row = i+1, column = 8).value == 'IN___SMS':	
			print "Incoming SMS from : " + sheet.cell(row = i+1, column = 1).value
		elif sheet.cell(row = i+1, column = 8).value == 'OUT___CALL':
			print "Outgoing Call to : " + sheet.cell(row = i+1, column = 2).value
		else:
			print "Outgoing SMS to : " + sheet.cell(row = i+1, column = 2).value

